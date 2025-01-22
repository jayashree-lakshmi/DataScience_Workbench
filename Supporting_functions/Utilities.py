import subprocess as sp
import itertools
from PyPDF2 import PdfFileReader, PdfFileWriter, utils
from collections import Counter, OrderedDict, Iterable
import os
import csv
import difflib
import pandas as pd
import openpyxl
from openpyxl import Workbook
import chardet
import json

def tiff2pdf(filename, new_filename):
    """
   :param filename: original tiff filename to be converted
   :param new_filename: converted filename to be saved with
    """

    try:
        file_extension = filename.split(".")[1].lower()

        if file_extension in ('tif','tiff'):
            print("Converting... .... TiFF file to PDF")

            cmd = [
                "tiff2pdf", "{}".format(filename), "-o", "{}".format(new_filename), "-p", "A4"
            ]

            result = sp.run(cmd)
            # result = sp.Popen(cmd, shell=True)
            # result.communicate()

            if result.returncode == 0:
                print(
                    "Converted {} to {} successfully".format(filename, new_filename))
            elif result.returncode == 2:
                print(
                    "PDF could not be genrated from {}".format(filename))
    except Exception as error:
        print(
            "Exception occured {} while converting tiff to pdf....Skipping tiff. Check if a valid tiff was downloaded"
            .format(error))

def split_pdf(filename, output_filename=None, start_range=None, end_range=None):
    """
    :param start_range: page number from where to start splitting
    :param end_range: page number from where to end splitting
    :return:
    """
    try:
        pdf_object = PdfFileReader(open(filename, "rb"))
        if pdf_object.isEncrypted:
            pdf_object.decrypt('')
  
        output = PdfFileWriter()
        for x in range(int(start_range) - 1, int(end_range)):
            output.addPage(pdf_object.getPage(x))
        if not output_filename:
            output_file = open("outfile" + str(end_range), "wb")
        else:
            output_file = open(output_filename, "wb")
        output.write(output_file)
        output_file.close()
        return output_filename
    except IndexError:
        print(
            "Please enter a valid start and end range for the PDF, it has only `{}` pages!!!"
            .format(pdf_object.getNumPages()))
    except Exception as error:
        print(error)

def check_patterns_df_columns(df_columnar_data, pattern):
    try:
        data = df_columnar_data.str.contains(pattern)
    except Exception as error:
        print(error)

def check_file_is_empty(filename):
    try:
        if os.stat(filename).st_size == 0:
            return True
        else:
            return False
    except Exception as error:
        print(error)

def flatten_list_list(list_of_list, unique=False):
    try:
        data = list(itertools.chain.from_iterable(list_of_list))
        if unique:
            data = list(set(data))
        return data
    except Exception as error:
        print(error)
    
def flatten_irregular_list(irregular_list):
    try:
        flattened_list = []
        for el in irregular_list:
            if isinstance(el, Iterable) and not isinstance(el, str):
                for sub in flatten_irregular_list(el):
                    flattened_list.append(sub)
            else:
                flattened_list.append(el)
        return flattened_list
    except Exception as error:
        print(error)

def write_csv_file(output_filename,
                   final_results,
                   write_mode='w',
                   delimiter=None,
                   quote_char=None,
                   quoting=None):
    """
    Writing content as a final output file as CSV
    :param output_filename:
    :return:
    """
    try:
        with open(output_filename, write_mode) as fp:
            print(
                "Writing results to CSV file: {} ".format(output_filename))
            if quote_char:
                quoting = csv.QUOTE_ALL
                delimiter = ","
                quotechar = quote_char
                writer = csv.writer(
                    fp,
                    delimiter=delimiter,
                    quotechar=quote_char,
                    quoting=quoting)
            else:
                writer = csv.writer(fp)
            writer.writerows(final_results)
    except IOError as ie:
        print(ie)
    except Exception as error:
        print(error)

def write_lists_csv(output_filename,
                    final_results,
                    write_mode='w',
                    delimiter=None,
                    quote_char=None,
                    quoting=None):

    # Write List of list to  CSV file
    try:
        with open(output_filename, 'w') as fp:
            writer = csv.writer(fp)
        writer.writerows(final_results)
        print("Writing results to CSV file: {} ".format(output_filename))
    except Exception as error:
        print(error)

def similarity_difflib(sent1, sent2):
    """
    Find similarity between two strings using difflib
    :param sent1: String1
    :param sent2: String 2
    :return: Ratio of matching
    """
    try:
        seq = difflib.SequenceMatcher(None, sent1, sent2)
        ratio = seq.ratio() * 100
        return ratio
    except Exception as error:
        print(error)

def write_dataframes_excel(pd, result, filename, sheet_name=None):
    """
    Write results to an excel sheet
    :param pd: Pandas object
    :param result: Results
    :param filename: Excel filename
    :param sheet_names: Excel sheetname
    :return: None
    """
    try:
        # write to excel
        writer = pd.ExcelWriter(filename)
        result.to_excel(writer, sheet_name)
        writer.save()
    except Exception as error:
        print(error)

def multiple_dfs_into_single_sheet(pd, df_list, sheets, file_name, spaces):
    # Usage:
    # list of dataframes
    # dfs = [df, df1, df2]
    #
    # # run function
    # multiple_dfs_into_single_sheet(dfs, 'Validation', 'test1.xlsx', 1)
    """
    Multiple dataframes pushed to a single sheet in excel workbook
    :param pd:
    :param df_list:
    :param sheets:
    :param file_name:
    :param spaces:
    :return:
    """

    try:
        writer = pd.ExcelWriter(file_name, engine='xlsxwriter')
        row = 0
        col = 0
        for dataframe in df_list:
            dataframe.to_excel(
                writer,
                sheet_name=sheets,
                startrow=row,
                startcol=col,
                index=False)

            row = row + len(dataframe.index) + spaces + 1
        writer.save()
    except Exception as error:
        print(error)

def dfs_different_tabs(pd, df_list, sheet_list, file_name):
    # Usage:
    # list of dataframes and sheet names
    # dfs = [df, df1, df2]
    # sheets = ['df', 'df1', 'df2']
    #
    # # run function
    # dfs_different_tabs(dfs, sheets, 'multi-test.xlsx')
    """
    Write multiple dataframes to different sheets in workbook
    :param pd:
    :param df_list:
    :param sheet_list:
    :param file_name:
    :return:
    """
    try:
        writer = pd.ExcelWriter(file_name, engine='xlsxwriter')
        for dataframe, sheet in zip(df_list, sheet_list):
            dataframe.to_excel(
                writer, sheet_name=sheet, startrow=0, startcol=0)
        writer.save()
    except Exception as error:
        print(error)

def detect_encoding_file(filename):
    """
    Detect File encoding type
    :param self:
    :param filename: Filename
    :return: Encoding Type(utf-8, ascii, iso-8859-1, latin1, cp1252)
    """
    try:
        with open(filename, 'rb') as f:
            result = chardet.detect(f.read())
        return result.get('encoding')
    except Exception as error:
        print(error)

def convert_df_to_json(filename, orient_type):
    data = pd.read_csv(filename).to_json(orient=orient_type)
    dataframe_json = json.loads(data)
    return dataframe_json


def combine_csv_to_xlsx(file_list, output_filename):
    """
    This function will consolidate csv to xlsx file
    :param file_list: list of csv files to be consolidated to xlsx
    :param output_filename: Filename of the output file
    usage: file_list = [('Roc-check.csv', 'Roc-check'), ('gst.csv', 'GST-Check')] ##<[(csv_filename, sheet_name)]
           output_filename = 'Financials.xlsx'
    """
    try:
        wb = Workbook()
        #file_list = [('Roc-check.csv', 'Roc-check'), ('gst.csv', 'GST-Check')]
        for index, item in enumerate(file_list):
            wb_write = wb.create_sheet(file_list[index][1])
            df = pd.read_csv(
                file_list[index][0],
                error_bad_lines=False,
                encoding="ISO-8859-1",
                delim_whitespace=False,
                quoting=0,
                delimiter=",")
            df = df.fillna(' ')
            df_list = df.values.tolist()
            df_header_list = df.columns.values.tolist()
            for i in range(len(df_header_list)):
                if 'Unnamed:' in df_header_list[i]:
                    df_header_list[i] = ' '
            df_list.insert(0, df_header_list)
            for row, value in enumerate(df_list):
                wb_write.append(value)
            wb.save(output_filename)
        wb_del = openpyxl.load_workbook(output_filename)
        sheet_del = wb_del.get_sheet_by_name('Sheet')
        wb_del.remove_sheet(sheet_del)
        wb_del.save(output_filename)

    except Exception as error:
        print(error)

def decode_password(pdf_file, password):
    """
    This function will decode a password protected file.
    :param pdf_file: input file path
    :param password: password required for opening the password protected file.
    """
    try:
        try:
            print("File: {} has been picked for password decryption.".
                        format(pdf_file))
            if not os.path.isfile(pdf_file) or os.path.splitext(
                    os.path.basename(pdf_file))[1] != '.pdf':

                print(
                    "The input file: {} either is not a pdf file or does not exist."
                    .format(os.path.basename(pdf_file)))
                return pdf_file, 'file_error'

        except utils.PdfReadError as err:
            print(err)
            return pdf_file, err.__str__()

        except NotImplementedError as err:
            print(err)
            return pdf_file, err.__str__()

        except Exception as error:
            print(error)
            return pdf_file, 'file_error'

        dest_file = "{}_decoded.pdf".format(
            os.path.splitext(os.path.basename(pdf_file))[0])
        with open(pdf_file, 'r') as input_file:
            with open(dest_file, 'w') as output_file:
                reader = PdfFileReader(input_file)
                try:
                    reader.decrypt(password)
                except KeyError:
                    return pdf_file, "not_password_protected"
                writer = PdfFileWriter()

                for i in range(reader.getNumPages()):
                    writer.addPage(reader.getPage(i))
                    writer.write(output_file)
            print("File has been successfully decoded to....{}".format(
                dest_file))
        pdf_file = 'decrypted_{}'.format(pdf_file)
        os.rename(dest_file, pdf_file)
        try:
            os.remove(dest_file)
        except:
            pass
        print("The password has been removed for `{}`".format(
            os.path.basename(pdf_file)))
        return pdf_file, "password_decrypted"

    except utils.PdfReadError as err:
        if err.__str__() == 'File has not been decrypted':
            return pdf_file, 'wrong_password'
        elif err.__str__() == 'EOF marker not found':
            return pdf_file, 'file_error'
        else:
            return pdf_file, err.__str__()
    except Exception as error:
        try:
            new_pdf_file = "decrypted_{}".format(pdf_file)
            res = sp.run([
                "qpdf", "--password={}".format(password), "--decrypt",
                "{}".format(pdf_file), new_pdf_file
            ])
            if res.returncode == 0:
                return new_pdf_file, "password_decrypted"
            elif res.returncode == 2:
                return pdf_file, "wrong_password"
            else:
                return pdf_file, "file_error"

        except Exception as error:
            print(" Failed, exception has occurred:  {}".format(error))
        return pdf_file, error.__str__()